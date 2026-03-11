from pathlib import Path

from openpyxl import load_workbook

import sys

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from database import init_db, normalize_username, upsert_channel  # noqa: E402

EXCEL_PATH = Path(__file__).resolve().parent / "tg_channels.xlsx"


def pick(row: dict, *names: str, default=""):
    for name in names:
        if name in row and row[name] is not None:
            return row[name]
    return default


def main():
    if not EXCEL_PATH.exists():
        print(f"❌ Excel не найден: {EXCEL_PATH}")
        return

    init_db()

    print(f"📥 Импорт из {EXCEL_PATH}...")

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("❌ Пустой Excel-файл")
        return

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    total = 0
    added = 0
    skipped = 0

    trusted_sources = {"tlgrm", "telegram.org.ru", "catalog_parser", "excel_import"}

    for values in rows[1:]:
        total += 1
        row = dict(zip(headers, values))

        username = normalize_username(
            str(
                pick(
                    row,
                    "username",
                    "Username",
                    "USERNAME",
                    "канал",
                    "channel",
                    "url",
                    default="",
                )
            )
        )

        if not username:
            skipped += 1
            continue

        title = str(pick(row, "title", "Title", "name", "Название", default="")).strip()
        description = str(
            pick(row, "description", "Description", "desc", "Описание", default="")
        ).strip()

        category = str(pick(row, "category", "Category", "Категория", default="")).strip()
        source = str(pick(row, "source", "Source", "Источник", default="excel_import")).strip()
        found_via = str(
            pick(row, "found_via", "Found_via", "Найдено_через", default="catalog_parser")
        ).strip()

        subscribers_raw = pick(row, "subscribers", "Subscribers", "Подписчики", default=0)
        try:
            subscribers = int(subscribers_raw or 0)
        except (TypeError, ValueError):
            subscribers = 0

        is_organization = 1 if (source in trusted_sources or found_via in trusted_sources) else 0

        upsert_channel(
            {
                "username": username,
                "title": title or username,
                "description": description,
                "subscribers": subscribers,
                "category": category,
                "source": source,
                "found_via": found_via,
                "is_organization": is_organization,
                "reviewed": 0,
            }
        )
        added += 1

        if added % 50 == 0:
            print(f"  ... {added} добавлено")

    print("✅ Импорт завершён!")
    print(f"   📊 Всего строк: {total}")
    print(f"   ✅ Добавлено/обновлено: {added}")
    print(f"   ⏭  Пропущено: {skipped}")


if __name__ == "__main__":
    main()