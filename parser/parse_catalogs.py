"""
Парсер каталогов Telegram-каналов → Excel

Источники:
1. tlgrm.ru — все категории
2. telegrator.ru — приоритетные категории с пагинацией

Приоритет: новости, экономика, наука, организации, IT, образование

Запуск:
    python parse_catalogs.py
"""

import requests
import re
import time
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
}

all_channels = {}  # username -> {title, subscribers, category, description, source}


# ═══════════════════════════════════════════════════
# TLGRM.RU
# ═══════════════════════════════════════════════════

# Приоритетные категории идут первыми
TLGRM_CATEGORIES = {
    # === ПРИОРИТЕТ ===
    "news":       "Медиа и СМИ",
    "economics":  "Бизнес и финансы",
    "business":   "Бизнес и финансы",
    "tech":       "IT и технологии",
    "education":  "Образование",
    "career":     "Бизнес и финансы",
    "sales":      "Ритейл и e-commerce",
    "health":     "Медицина",
    "books":      "Образование",
    "marketing":  "Бизнес и финансы",
    # === ОСТАЛЬНЫЕ ===
    "crypto":     "Бизнес и финансы",
    "travel":     "Другое",
    "music":      "Другое",
    "food":       "Другое",
    "psychology": "Другое",
    "blogs":      "Другое",
    "video":      "Другое",
    "art":        "Другое",
    "quotes":     "Другое",
    "other":      "Другое",
}


def parse_tlgrm_category(category_slug):
    url = f"https://tlgrm.ru/channels/{category_slug}"
    category_name = TLGRM_CATEGORIES.get(category_slug, "Другое")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            print(f"[!] HTTP {resp.status_code}")
            return 0
        soup = BeautifulSoup(resp.text, "html.parser")
        count = 0
        for link in soup.find_all("a", href=True):
            href = link.get("href", "")
            m = re.search(r"/channels/@([a-zA-Z_][a-zA-Z0-9_]+)", href)
            if not m:
                continue
            username = m.group(1)
            if username in all_channels:
                continue
            parent = link.find_parent("div") or link.find_parent("li") or link.parent
            title_el = link.find("h3") or link.find("strong") or link
            title = title_el.get_text(strip=True) if title_el else username
            subs = 0
            text_around = parent.get_text() if parent else ""
            subs_match = re.findall(r"([\d,]{4,})", text_around.replace(" ", ""))
            for s in subs_match:
                try:
                    n = int(s.replace(",", ""))
                    if n > subs and n < 100_000_000:
                        subs = n
                except:
                    pass
            desc = ""
            for p in (parent.find_all("p") if parent else []):
                t = p.get_text(strip=True)
                if t and len(t) > 10 and username not in t:
                    desc = t[:300]
                    break
            all_channels[username] = {
                "title": title,
                "subscribers": subs,
                "category": category_name,
                "description": desc,
                "source": "tlgrm.ru",
            }
            count += 1
        return count
    except Exception as e:
        print(f"[!] Ошибка: {e}")
        return 0


def parse_tlgrm():
    print("\n" + "=" * 60)
    print("📡 [1/2] Парсим tlgrm.ru")
    print("=" * 60)
    total = 0
    for slug in TLGRM_CATEGORIES:
        print(f"  [{slug}]", end=" ")
        count = parse_tlgrm_category(slug)
        total += count
        print(f"→ +{count} (всего: {len(all_channels)})")
        time.sleep(1)
    print(f"\n✅ tlgrm.ru: +{total}")


# ═══════════════════════════════════════════════════
# TELEGRATOR.RU (с пагинацией!)
# ═══════════════════════════════════════════════════

# Приоритетные категории — больше страниц парсим
TELEGRATOR_CATEGORIES = {
    # slug: (наша_категория, сколько_страниц)
    # === ПРИОРИТЕТ — берём больше ===
    "novosti":    ("Медиа и СМИ", 10),
    "biznes":     ("Бизнес и финансы", 10),
    "nauka":      ("Наука и исследования", 10),
    "it":         ("IT и технологии", 8),
    "politika":   ("Госорганы", 5),
    "investing":  ("Бизнес и финансы", 5),
    "rabota-vakansii": ("Бизнес и финансы", 5),
    "marketing":  ("Маркетинг и PR", 5),
    # === ОСТАЛЬНЫЕ — по 3 страницы ===
    "psihologiya": ("Образование", 3),
    "lingvistika": ("Образование", 3),
    "seo":        ("IT и технологии", 3),
    "avto":       ("Другое", 2),
    "krasota-i-zdorove": ("Медицина", 3),
    "puteshestviya": ("Другое", 2),
    "kulinariya": ("Другое", 2),
    "blogi":      ("Другое", 2),
    "design":     ("Другое", 2),
    "skidki":     ("Ритейл и e-commerce", 3),
    "magazine":   ("Ритейл и e-commerce", 3),
}


def parse_telegrator_page(category_slug, category_name, page=1):
    """Парсить одну страницу категории telegrator.ru."""
    url = f"https://telegrator.ru/channels/{category_slug}/"
    if page > 1:
        url = f"https://telegrator.ru/channels/{category_slug}/page/{page}/"

    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return 0, False  # 0 каналов, нет следующей страницы

        soup = BeautifulSoup(resp.text, "html.parser")
        count = 0

        # Ищем ссылки на t.me/username
        for link in soup.find_all("a", href=True):
            href = link.get("href", "")
            m = re.search(r"t\.me/([a-zA-Z_][a-zA-Z0-9_]{3,30})", href)
            if not m:
                continue

            username = m.group(1)
            if username in all_channels:
                continue
            if username.lower() in ("joinchat", "addstickers", "share", "proxy", "socks"):
                continue

            # Ищем контейнер карточки
            card = link.find_parent("article") or link.find_parent("div", class_=True)

            # Название
            title = ""
            if card:
                h2 = card.find("h2") or card.find("h3")
                if h2:
                    title = h2.get_text(strip=True)
            if not title:
                title = link.get_text(strip=True) or username

            # Подписчики
            subs = 0
            if card:
                text = card.get_text()
                nums = re.findall(r"([\d\s]{4,})", text)
                for n in nums:
                    try:
                        val = int(n.replace(" ", "").replace("\xa0", ""))
                        if val > subs and val < 100_000_000:
                            subs = val
                    except:
                        pass

            # Описание
            desc = ""
            if card:
                for p in card.find_all(["p", "div"]):
                    t = p.get_text(strip=True)
                    if t and len(t) > 20 and username.lower() not in t.lower() and "подписчик" not in t.lower():
                        desc = t[:300]
                        break

            all_channels[username] = {
                "title": title,
                "subscribers": subs,
                "category": category_name,
                "description": desc,
                "source": "telegrator.ru",
            }
            count += 1

        # Проверяем есть ли следующая страница
        has_next = bool(soup.find("a", href=re.compile(rf"/page/{page + 1}/")))
        return count, has_next

    except Exception as e:
        print(f"[!] Ошибка: {e}")
        return 0, False


def parse_telegrator():
    print("\n" + "=" * 60)
    print("📡 [2/2] Парсим telegrator.ru (с пагинацией)")
    print("=" * 60)

    total = 0
    for slug, (category_name, max_pages) in TELEGRATOR_CATEGORIES.items():
        print(f"  [{slug}]", end=" ")
        cat_count = 0

        for page in range(1, max_pages + 1):
            count, has_next = parse_telegrator_page(slug, category_name, page)
            cat_count += count
            time.sleep(1.5)

            if not has_next or count == 0:
                break

        total += cat_count
        print(f"→ +{cat_count} (всего: {len(all_channels)})")

    print(f"\n✅ telegrator.ru: +{total}")


# ═══════════════════════════════════════════════════
# СОХРАНЕНИЕ В EXCEL
# ═══════════════════════════════════════════════════

# Приоритетные категории для сортировки
PRIORITY_CATEGORIES = [
    "Медиа и СМИ",
    "Бизнес и финансы",
    "Наука и исследования",
    "IT и технологии",
    "Госорганы",
    "Образование",
    "Маркетинг и PR",
    "Медицина",
    "Ритейл и e-commerce",
]


def category_sort_key(item):
    """Приоритетные категории первыми, потом по подписчикам."""
    username, data = item
    cat = data["category"]
    if cat in PRIORITY_CATEGORIES:
        priority = PRIORITY_CATEGORIES.index(cat)
    else:
        priority = 100
    return (priority, -data["subscribers"])


def save_to_excel(filepath="tg_channels.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Каналы"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    priority_fill = PatternFill("solid", fgColor="E8F5E9")
    normal_fill = PatternFill("solid", fgColor="F5F5F5")

    headers = ["#", "Username", "Название", "Подписчики", "Категория", "Описание", "Источник"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sorted_channels = sorted(all_channels.items(), key=category_sort_key)

    for i, (username, data) in enumerate(sorted_channels, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=f"@{username}").border = thin_border
        ws.cell(row=row, column=3, value=data["title"]).border = thin_border
        ws.cell(row=row, column=4, value=data["subscribers"]).border = thin_border
        ws.cell(row=row, column=5, value=data["category"]).border = thin_border
        ws.cell(row=row, column=6, value=data["description"]).border = thin_border
        ws.cell(row=row, column=7, value=data["source"]).border = thin_border

        is_priority = data["category"] in PRIORITY_CATEGORIES
        fill = priority_fill if is_priority else (normal_fill if i % 2 == 0 else None)
        if fill:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 18

    ws.auto_filter.ref = f"A1:G{len(sorted_channels) + 1}"
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"\n💾 Сохранено: {filepath} ({len(sorted_channels)} каналов)")


# ═══════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("🔍 Парсер каталогов Telegram-каналов → Excel")
    print("   Приоритет: новости, экономика, наука, организации")
    print("=" * 60)

    parse_tlgrm()
    parse_telegrator()

    print(f"\n{'=' * 60}")
    print(f"📊 ИТОГО: {len(all_channels)} уникальных каналов")
    print(f"{'=' * 60}")

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tg_channels.xlsx")
    save_to_excel(output_path)

    # Статистика
    cats = {}
    sources = {}
    for ch in all_channels.values():
        cats[ch["category"]] = cats.get(ch["category"], 0) + 1
        sources[ch["source"]] = sources.get(ch["source"], 0) + 1

    print("\n📂 По категориям:")
    for cat, cnt in sorted(cats.items(), key=lambda x: -x[1]):
        marker = "⭐" if cat in PRIORITY_CATEGORIES else "  "
        print(f"   {marker} {cat}: {cnt}")

    print("\n📡 По источникам:")
    for src, cnt in sorted(sources.items(), key=lambda x: -x[1]):
        print(f"   {src}: {cnt}")


if __name__ == "__main__":
    main()